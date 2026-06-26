from __future__ import annotations

import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from copy import deepcopy
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl


CASHBACK_RATE = 0.06
DEFAULT_TRACKER_EXPORT = Path("C:/Users/coope/Downloads/tracker_data (6).xlsx")
DEFAULT_PRICE_WORKBOOK = Path("C:/Users/coope/Downloads/BFMR Tracking.xlsx")
ORDER_NUMBER_PATTERN = re.compile(r"\b\d{3}-\d{7}-\d{7}\b")
MANUAL_ASSUMED_ORDERS = {"111-1403104-8336261"}
DEFAULT_SETTINGS: dict[str, Any] = {
    "assumptions": {
        "default_cashback_rate": CASHBACK_RATE,
        "no_order_account": "Personal",
        "no_order_cashback_rate": CASHBACK_RATE,
        "business_default_cashback_rate": CASHBACK_RATE,
        "manual_assumed_orders": [
            {
                "order_number": "111-1403104-8336261",
                "account": "Personal",
                "cashback_rate": CASHBACK_RATE,
                "note": "Manual 6% assumption per user",
            }
        ],
    },
    "chrome": {
        "bfmr_profile_directory": "Default",
        "skip_paid_orders": True,
        "profiles": [
            {
                "id": "personal-default",
                "name": "Personal Amazon",
                "profile_directory": "Default",
                "account_type": "personal",
                "enabled": True,
            },
            {
                "id": "business-profile-9",
                "name": "Business Amazon",
                "profile_directory": "Profile9",
                "account_type": "business",
                "enabled": True,
            },
        ],
    },
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return ""
        if float(value).is_integer():
            return f"{value:.0f}"
        return str(value)
    text = str(value).replace("\xa0", " ").replace("\u200e", "")
    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"\s+", " ", text).strip()
    if re.fullmatch(r"\d+(\.\d+)?e\+\d+", text, flags=re.IGNORECASE):
        try:
            return format(Decimal(text), "f").split(".")[0]
        except InvalidOperation:
            return text
    return text


def normalize_text(value: Any) -> str:
    return clean_text(value).lower()


def item_base(value: Any) -> str:
    text = normalize_text(value)
    return re.sub(r"\s+-\s+[a-z0-9]{5,}(/[a-z])?$", "", text)


def parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = clean_text(value).replace("$", "").replace(",", "")
    if text in {"", "-"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        # Excel serial date, using the common 1899-12-30 base.
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()

    text = clean_text(value)
    text = re.sub(r"\s+(ET|EST|EDT|CT|CST|CDT|PT|PST|PDT)$", "", text)
    for fmt in ("%m-%d-%Y %I:%M %p", "%m/%d/%Y %I:%M %p", "%m-%d-%Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def parse_amazon_eta_date(eta: Any, scraped_at: Any = None) -> str:
    value = clean_text(eta)
    if not value:
        return ""

    base = parse_date(scraped_at) or date.today().isoformat()
    base_date = date.fromisoformat(base)
    lowered = value.lower()
    if re.search(r"\btoday\b", lowered):
        return base_date.isoformat()
    if re.search(r"\btomorrow\b", lowered):
        return (base_date + timedelta(days=1)).isoformat()

    month_names = {
        "jan": 1,
        "january": 1,
        "feb": 2,
        "february": 2,
        "mar": 3,
        "march": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "jun": 6,
        "june": 6,
        "jul": 7,
        "july": 7,
        "aug": 8,
        "august": 8,
        "sep": 9,
        "sept": 9,
        "september": 9,
        "oct": 10,
        "october": 10,
        "nov": 11,
        "november": 11,
        "dec": 12,
        "december": 12,
    }
    month_match = re.search(
        r"\b("
        r"jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|"
        r"sep(?:t|tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?"
        r")\s+(\d{1,2})\b",
        lowered,
    )
    if month_match:
        month = month_names[month_match.group(1)]
        day = int(month_match.group(2))
        candidate = date(base_date.year, month, day)
        if candidate < base_date - timedelta(days=7):
            candidate = date(base_date.year + 1, month, day)
        return candidate.isoformat()

    weekdays = {
        "monday": 0,
        "mon": 0,
        "tuesday": 1,
        "tue": 1,
        "wednesday": 2,
        "wed": 2,
        "thursday": 3,
        "thu": 3,
        "friday": 4,
        "fri": 4,
        "saturday": 5,
        "sat": 5,
        "sunday": 6,
        "sun": 6,
    }
    weekday_match = re.search(r"\b(monday|mon|tuesday|tue|wednesday|wed|thursday|thu|friday|fri|saturday|sat|sunday|sun)\b", lowered)
    if weekday_match:
        target = weekdays[weekday_match.group(1)]
        days_ahead = (target - base_date.weekday()) % 7
        return (base_date + timedelta(days=days_ahead)).isoformat()

    parsed = parse_date(value)
    return parsed or ""


def meaningful_amazon_eta(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if parse_amazon_eta_date(text):
        return text
    return ""


def header_map(ws: Any, row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row, col).value
        if value not in (None, ""):
            mapping[clean_text(value)] = col
    return mapping


def find_header_row(ws: Any, required: set[str], max_rows: int = 15) -> int:
    wanted = {name.lower() for name in required}
    for row in range(1, min(ws.max_row, max_rows) + 1):
        headers = {clean_text(ws.cell(row, col).value).lower() for col in range(1, ws.max_column + 1)}
        if wanted.issubset(headers):
            return row
    raise ValueError(f"Could not find workbook headers: {', '.join(sorted(required))}")


def cell(ws: Any, headers: dict[str, int], row: int, name: str) -> Any:
    col = headers.get(name)
    if not col:
        return None
    return ws.cell(row, col).value


def account_from_order(order_number: str) -> str:
    if order_number:
        return "Amazon unmatched"
    return "Unknown"


def normalize_order_number(value: Any) -> str:
    text = clean_text(value)
    match = ORDER_NUMBER_PATTERN.search(text)
    return match.group(0) if match else text


def has_amazon_order_number(value: Any) -> bool:
    return bool(ORDER_NUMBER_PATTERN.fullmatch(normalize_order_number(value)))


def merged_settings(settings: dict[str, Any] | None = None) -> dict[str, Any]:
    merged = deepcopy(DEFAULT_SETTINGS)
    if not isinstance(settings, dict):
        return merged
    for section, values in settings.items():
        if isinstance(values, dict) and isinstance(merged.get(section), dict):
            merged[section].update(values)
        else:
            merged[section] = values
    return merged


def assumption_float(settings: dict[str, Any], key: str, fallback: float = CASHBACK_RATE) -> float:
    try:
        return float(settings.get("assumptions", {}).get(key, fallback))
    except (TypeError, ValueError):
        return fallback


def manual_assumption_for_order(settings: dict[str, Any], order_number: str) -> dict[str, Any] | None:
    normalized = normalize_order_number(order_number)
    for row in settings.get("assumptions", {}).get("manual_assumed_orders", []):
        if normalize_order_number(row.get("order_number")) == normalized:
            return row
    if normalized in MANUAL_ASSUMED_ORDERS:
        return {
            "order_number": normalized,
            "account": "Personal",
            "cashback_rate": CASHBACK_RATE,
            "note": "Manual 6% assumption per user",
        }
    return None


def calculate_profit(status: str, payout_total: float, purchase_total: float, cashback_rate: float) -> float:
    if status.lower() == "cancelled":
        return 0.0
    return payout_total - purchase_total + (purchase_total * cashback_rate)


def is_referral_bonus_item(value: Any) -> bool:
    return normalize_text(value) == "referral bonus"


def parse_cashback_rate(text: Any, fallback: float | None = None) -> tuple[float | None, str]:
    value = clean_text(text)
    percents = [float(match) for match in re.findall(r"(\d+(?:\.\d+)?)\s*%", value)]
    if not percents:
        return fallback, f"Default {fallback:.0%}" if fallback is not None else ""
    rate = sum(percents) / 100
    if rate <= 0 or rate > 0.25:
        return fallback, f"Default {fallback:.0%}" if fallback is not None else "Unparsed"
    return rate, value


def normalize_amazon_order(
    row: dict[str, Any],
    account: str,
    profile_label: str,
    source_url: str = "",
    settings: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    settings = merged_settings(settings)
    default_rate = assumption_float(settings, "default_cashback_rate")
    business_rate = assumption_float(settings, "business_default_cashback_rate", default_rate)
    order_number = normalize_order_number(
        first_present(row, ["order_number", "Order #", "Order", "Order Number", "Order ID", "orderId"])
    )
    if not ORDER_NUMBER_PATTERN.fullmatch(order_number):
        return None

    reward_text = clean_text(
        first_present(row, ["reward_text", "Rewards", "Cashback", "Cash Back", "Payment Rewards", "Payment method"])
    )
    payment_method = clean_text(first_present(row, ["payment_method", "Payment Method", "Payment", "Card"]))
    account_label = "Business" if account.lower() == "business" else "Personal"
    explicit_rate = parse_number(first_present(row, ["cashback_rate", "Cashback Rate", "Cash Back Rate"]))
    if explicit_rate and explicit_rate > 1:
        explicit_rate /= 100

    if account_label == "Business":
        rate = explicit_rate if explicit_rate is not None else business_rate
        rate_source = f"Business default {business_rate:.0%}"
        if reward_text:
            rate_source = f"Business default {business_rate:.0%}; Amazon displayed: {reward_text}"
    elif explicit_rate is not None:
        rate = explicit_rate
        rate_source = reward_text or "Personal Amazon order history"
    else:
        rate, rate_source = parse_cashback_rate(reward_text, default_rate)
        if rate_source.startswith("Default"):
            rate_source = f"Personal default {default_rate:.0%} pending visible reward text"

    delivery_status = clean_text(first_present(row, ["delivery_status", "Delivery Status", "Shipment Status"]))
    delivery_eta = meaningful_amazon_eta(first_present(row, ["delivery_eta", "Delivery ETA", "ETA", "Arriving"]))
    delivery_scraped_at = clean_text(first_present(row, ["delivery_scraped_at", "Delivery Scraped At"]))
    delivery_eta_date = (
        parse_date(first_present(row, ["delivery_eta_date", "Delivery ETA Date"]))
        or parse_amazon_eta_date(delivery_eta, delivery_scraped_at)
        or parse_amazon_eta_date(delivery_status, delivery_scraped_at)
    )
    if not delivery_eta and delivery_eta_date:
        delivery_eta = delivery_eta_date

    return {
        "order_number": order_number,
        "account": account_label,
        "profile": clean_text(profile_label),
        "cashback_rate": round(float(rate or default_rate), 4),
        "cashback_rate_source": rate_source,
        "payment_method": payment_method,
        "reward_text": reward_text,
        "order_date": parse_date(first_present(row, ["order_date", "Order Date", "Date"])) or "",
        "order_total": parse_number(first_present(row, ["order_total", "Order Total", "Total"])),
        "source_url": source_url,
        "detail_url": clean_text(first_present(row, ["detail_url", "Detail URL", "Order Detail URL"])),
        "delivery_status": delivery_status,
        "delivery_eta": delivery_eta,
        "delivery_eta_date": delivery_eta_date,
        "delivery_scraped_at": delivery_scraped_at,
        "raw_text": clean_text(first_present(row, ["raw_text", "Raw Text", "Page Text"]))[:8000],
        "line_items": first_present(row, ["line_items", "Line Items"]) or [],
        "detail_scraped_at": clean_text(first_present(row, ["detail_scraped_at", "Detail Scraped At"])),
        "scraped_at": datetime.now().isoformat(timespec="seconds"),
    }


def normalize_amazon_orders(
    rows: list[dict[str, Any]],
    account: str,
    profile_label: str,
    source_url: str = "",
    settings: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows:
        if not isinstance(row, dict):
            continue
        order = normalize_amazon_order(row, account, profile_label, source_url, settings)
        if not order or order["order_number"] in seen:
            continue
        seen.add(order["order_number"])
        normalized.append(order)
    return normalized


def apply_amazon_enrichment(
    dataset: dict[str, Any],
    amazon_orders: list[dict[str, Any]],
    settings: dict[str, Any] | None = None,
) -> dict[str, Any]:
    settings = merged_settings(settings)
    default_rate = assumption_float(settings, "default_cashback_rate")
    no_order_rate = assumption_float(settings, "no_order_cashback_rate", default_rate)
    no_order_account = clean_text(settings.get("assumptions", {}).get("no_order_account")) or "Personal"
    enriched = deepcopy(dataset)
    infer_partial_return_details(enriched.get("records", []))
    infer_split_return_details(enriched.get("records", []))
    orders_by_number = {
        normalize_order_number(order.get("order_number")): order
        for order in amazon_orders
        if normalize_order_number(order.get("order_number"))
    }
    matched = 0
    personal = 0
    business = 0
    precise_cashback = 0

    for record in enriched.get("records", []):
        order_number = normalize_order_number(record.get("order_number"))
        if clean_text(record.get("status")).lower() == "cancelled":
            record["account"] = "Ignored cancelled"
            record["account_source"] = "BFMR status is cancelled"
            record["amazon_order_matched"] = False
            record["amazon_profile"] = ""
            record["amazon_payment_method"] = ""
            record["amazon_reward_text"] = ""
            record["amazon_delivery_status"] = ""
            record["amazon_delivery_eta"] = ""
            record["amazon_delivery_eta_date"] = ""
            record["amazon_delivery_scraped_at"] = ""
            record["cashback_rate"] = default_rate
            record["cashback_rate_source"] = "Ignored because BFMR status is cancelled"
            record["profit"] = 0.0
            continue
        if is_referral_bonus_item(record.get("item_name")):
            record["account"] = "BFMR Referral"
            record["account_source"] = "BFMR payment history referral bonus"
            record["amazon_order_matched"] = False
            record["amazon_profile"] = ""
            record["amazon_payment_method"] = ""
            record["amazon_reward_text"] = ""
            record["amazon_delivery_status"] = ""
            record["amazon_delivery_eta"] = ""
            record["amazon_delivery_eta_date"] = ""
            record["amazon_delivery_scraped_at"] = ""
            record["cashback_rate"] = 0.0
            record["cashback_rate_source"] = "No cashback on BFMR referral bonus"
            record["profit"] = round(float(record.get("payout_total") or 0), 2)
            continue
        manual_assumption = manual_assumption_for_order(settings, order_number)
        if manual_assumption:
            manual_rate = parse_number(manual_assumption.get("cashback_rate"))
            manual_rate = float(manual_rate if manual_rate is not None else default_rate)
            manual_account = clean_text(manual_assumption.get("account")) or "Personal"
            manual_note = clean_text(manual_assumption.get("note")) or f"Manual {manual_rate:.0%} assumption"
            matched += 1
            if manual_account == "Personal":
                personal += 1
            if manual_account == "Business":
                business += 1
            record["account"] = manual_account
            record["account_source"] = manual_note
            record["amazon_order_matched"] = True
            record["amazon_profile"] = ""
            record["amazon_payment_method"] = ""
            record["amazon_reward_text"] = ""
            record["amazon_delivery_status"] = ""
            record["amazon_delivery_eta"] = ""
            record["amazon_delivery_eta_date"] = ""
            record["amazon_delivery_scraped_at"] = ""
            record["cashback_rate"] = round(manual_rate, 4)
            record["cashback_rate_source"] = manual_note
            record["profit"] = round(
                calculate_profit(
                    record.get("status", ""),
                    float(record.get("payout_total") or 0),
                    float(record.get("purchase_total") or 0),
                    manual_rate,
                ),
                2,
            )
            continue
        amazon = orders_by_number.get(order_number)
        if amazon:
            matched += 1
            account = clean_text(amazon.get("account")) or "Amazon matched"
            if account == "Personal":
                personal += 1
            if account == "Business":
                business += 1
            rate = float(amazon.get("cashback_rate") or default_rate)
            source = clean_text(amazon.get("cashback_rate_source")) or "Amazon order history"
            if "default" not in source.lower():
                precise_cashback += 1
            record["account"] = account
            record["account_source"] = f"Amazon {account} order history"
            record["amazon_order_matched"] = True
            record["amazon_profile"] = clean_text(amazon.get("profile"))
            record["amazon_payment_method"] = clean_text(amazon.get("payment_method"))
            record["amazon_reward_text"] = clean_text(amazon.get("reward_text"))
            record["amazon_delivery_status"] = clean_text(amazon.get("delivery_status"))
            record["amazon_delivery_eta"] = meaningful_amazon_eta(amazon.get("delivery_eta"))
            record["amazon_delivery_eta_date"] = clean_text(amazon.get("delivery_eta_date")) or parse_amazon_eta_date(
                record["amazon_delivery_eta"] or amazon.get("delivery_status"), amazon.get("delivery_scraped_at")
            )
            if not record["amazon_delivery_eta"] and record["amazon_delivery_eta_date"]:
                record["amazon_delivery_eta"] = record["amazon_delivery_eta_date"]
            record["amazon_delivery_scraped_at"] = clean_text(amazon.get("delivery_scraped_at"))
            record["cashback_rate"] = round(rate, 4)
            record["cashback_rate_source"] = source
        else:
            record["account"] = "Amazon unmatched" if order_number else no_order_account
            record["account_source"] = (
                "Waiting for Amazon order history match"
                if order_number
                else f"No Amazon order number; assumed {no_order_account} at {no_order_rate:.0%}"
            )
            record["amazon_order_matched"] = False
            record["amazon_profile"] = ""
            record["amazon_payment_method"] = ""
            record["amazon_reward_text"] = ""
            record["amazon_delivery_status"] = ""
            record["amazon_delivery_eta"] = ""
            record["amazon_delivery_eta_date"] = ""
            record["amazon_delivery_scraped_at"] = ""
            record["cashback_rate"] = default_rate if order_number else no_order_rate
            record["cashback_rate_source"] = (
                f"Default {default_rate:.0%} pending Amazon match" if order_number else f"No-order default {no_order_rate:.0%}"
            )

        record["profit"] = round(
            calculate_profit(
                record.get("status", ""),
                float(record.get("payout_total") or 0),
                float(record.get("purchase_total") or 0),
                float(record.get("cashback_rate") or CASHBACK_RATE),
            ),
            2,
        )

    enriched["summary"] = summarize(enriched.get("records", []))
    metadata = enriched.setdefault("metadata", {})
    active_amazon_records = [
        record
        for record in enriched.get("records", [])
        if clean_text(record.get("status")).lower() != "cancelled"
        and has_amazon_order_number(record.get("order_number"))
    ]
    metadata["amazon_order_count"] = len(amazon_orders)
    metadata["amazon_matched_orders"] = matched
    metadata["amazon_unmatched_orders"] = max(len(active_amazon_records) - matched, 0)
    metadata["amazon_personal_matches"] = personal
    metadata["amazon_business_matches"] = business
    metadata["amazon_precise_cashback_matches"] = precise_cashback
    metadata["amazon_enriched_at"] = datetime.now().isoformat(timespec="seconds")
    return enriched


def display_status(status: str) -> str:
    value = clean_text(status)
    if not value:
        return "Unknown"
    return value[:1].upper() + value[1:].lower()


def load_price_lookup(price_workbook: Path | None) -> dict[str, Any]:
    lookup: dict[str, Any] = {
        "exact": {},
        "base_order": {},
        "order": defaultdict(list),
    }
    if not price_workbook or not Path(price_workbook).exists():
        return lookup

    wb = openpyxl.load_workbook(price_workbook, data_only=True)
    ws = wb.active
    row = find_header_row(ws, {"ITEM NAME", "ORDER #", "PURCHASE $"})
    headers = header_map(ws, row)

    for r in range(row + 1, ws.max_row + 1):
        item = cell(ws, headers, r, "ITEM NAME")
        order = clean_text(cell(ws, headers, r, "ORDER #"))
        purchase = parse_number(cell(ws, headers, r, "PURCHASE $"))
        if not item or not order or purchase is None or purchase <= 0:
            continue
        record = {
            "purchase_total": purchase,
            "item": clean_text(item),
            "order": order,
        }
        lookup["exact"][(normalize_text(item), order)] = record
        lookup["base_order"][(item_base(item), order)] = record
        lookup["order"][order].append(record)
    return lookup


def resolve_purchase_total(row_values: dict[str, Any], price_lookup: dict[str, Any]) -> tuple[float, str, bool]:
    item = row_values.get("Items")
    if is_referral_bonus_item(item):
        return 0.0, "Referral bonus", False

    retail = parse_number(row_values.get("Retail Price"))
    if retail is not None and retail > 0:
        return retail, "BFMR Retail Price", False

    order = clean_text(row_values.get("Order No."))
    exact = price_lookup["exact"].get((normalize_text(item), order))
    if exact:
        return exact["purchase_total"], "Existing tracker match", False

    base_match = price_lookup["base_order"].get((item_base(item), order))
    if base_match:
        return base_match["purchase_total"], "Existing tracker match", False

    order_matches = price_lookup["order"].get(order, [])
    if order and len(order_matches) == 1:
        return order_matches[0]["purchase_total"], "Existing tracker order match", False

    payout_total = parse_number(row_values.get("Subtotal")) or 0.0
    return payout_total, "Payout fallback", True


def first_present(row: dict[str, Any], names: list[str]) -> Any:
    normalized = {normalize_text(key): value for key, value in row.items()}
    for name in names:
        key = normalize_text(name)
        if key in normalized and normalized[key] not in (None, ""):
            return normalized[key]
    return None


def infer_site_purchase_total(
    item: Any, retail_price: float | None, quantity: float, payout_total: float
) -> tuple[float, str, bool]:
    if is_referral_bonus_item(item):
        return 0.0, "Referral bonus", False
    if retail_price is None or retail_price <= 0:
        return payout_total, "Payout fallback", True
    if quantity > 1 and payout_total and retail_price < payout_total * 0.75:
        return retail_price * quantity, "BFMR site retail price", False
    return retail_price, "BFMR site retail price", False


def infer_partial_return_details(records: list[dict[str, Any]]) -> None:
    candidates_by_item: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        status = clean_text(record.get("status")).lower()
        order_number = normalize_order_number(record.get("order_number"))
        if status == "cancelled" or not has_amazon_order_number(order_number):
            continue
        candidates_by_item[normalize_text(record.get("item_name"))].append(record)

    for record in records:
        status = clean_text(record.get("status")).lower()
        if status != "return" or has_amazon_order_number(record.get("order_number")):
            continue

        candidates = candidates_by_item.get(normalize_text(record.get("item_name")), [])
        if not candidates:
            continue

        if record.get("date"):
            dated_candidates = [
                candidate
                for candidate in candidates
                if candidate.get("date")
                and abs((date.fromisoformat(record["date"]) - date.fromisoformat(candidate["date"])).days) <= 14
            ]
            if dated_candidates:
                candidates = dated_candidates

        order_numbers = {normalize_order_number(candidate.get("order_number")) for candidate in candidates}
        order_numbers.discard("")
        if len(order_numbers) != 1:
            continue

        inferred_order = next(iter(order_numbers))
        record["order_number"] = inferred_order
        record["account"] = account_from_order(inferred_order)
        record["account_source"] = "Inferred from same-item partial return"
        record["order_number_inferred"] = True

        purchase_values = {
            round(float(candidate.get("purchase_total") or 0), 2)
            for candidate in candidates
            if float(candidate.get("purchase_total") or 0) > 0
        }
        if record.get("purchase_is_estimate") and len(purchase_values) == 1:
            record["purchase_total"] = next(iter(purchase_values))
            record["purchase_is_estimate"] = False
            record["price_source"] = "Inferred from same-item partial return"
            record["profit"] = round(
                calculate_profit(
                    clean_text(record.get("status")),
                    float(record.get("payout_total") or 0),
                    float(record.get("purchase_total") or 0),
                    float(record.get("cashback_rate") or CASHBACK_RATE),
                ),
                2,
            )


def infer_split_return_details(records: list[dict[str, Any]]) -> None:
    candidates_by_item: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        status = clean_text(record.get("status")).lower()
        order_number = normalize_order_number(record.get("order_number"))
        if status == "cancelled" or not has_amazon_order_number(order_number):
            continue
        candidates_by_item[normalize_text(record.get("item_name"))].append(record)

    for record in records:
        record.pop("split_review_needed", None)
        record.pop("split_review_reason", None)
        record.pop("split_candidate_orders", None)
        record.pop("return_group_key", None)
        record.pop("return_context", None)

    for record in records:
        status = clean_text(record.get("status")).lower()
        if status == "cancelled" or is_referral_bonus_item(record.get("item_name")):
            continue
        item_key = normalize_text(record.get("item_name"))
        candidates = [candidate for candidate in candidates_by_item.get(item_key, []) if candidate is not record]
        if not candidates:
            continue

        if has_amazon_order_number(record.get("order_number")):
            order_number = normalize_order_number(record.get("order_number"))
            related = [candidate for candidate in candidates if normalize_order_number(candidate.get("order_number")) == order_number]
            related_statuses = {clean_text(candidate.get("status")).lower() for candidate in related}
            if status in {"return", "deadline"} or related_statuses.intersection({"return", "deadline"}):
                record["return_group_key"] = order_number
                record["return_context"] = "Return/split group with matching same-item order rows"
            continue

        record_date = parse_date(record.get("date"))
        dated_candidates = candidates
        if record_date:
            row_date = date.fromisoformat(record_date)
            dated_candidates = []
            for candidate in candidates:
                candidate_date = parse_date(candidate.get("date"))
                if not candidate_date:
                    continue
                days = (row_date - date.fromisoformat(candidate_date)).days
                if -1 <= days <= 21:
                    dated_candidates.append(candidate)
            if not dated_candidates:
                dated_candidates = candidates

        order_numbers = sorted(
            {
                normalize_order_number(candidate.get("order_number"))
                for candidate in dated_candidates
                if has_amazon_order_number(candidate.get("order_number"))
            }
        )
        if len(order_numbers) == 1:
            inferred_order = order_numbers[0]
            record["order_number"] = inferred_order
            record["order_number_inferred"] = True
            record["return_group_key"] = inferred_order
            record["return_context"] = "Inferred from same-item split/return row"
            record["account"] = account_from_order(inferred_order)
            record["account_source"] = "Inferred from same-item split/return row"
        elif len(order_numbers) > 1:
            record["split_review_needed"] = True
            record["split_review_reason"] = "Multiple same-item Amazon orders are plausible; original order was not inferred"
            record["split_candidate_orders"] = order_numbers
            record["return_context"] = "Needs manual original-order review"


def normalize_bfmr_site_rows(rows: list[dict[str, Any]], source_url: str = "") -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        item = clean_text(first_present(row, ["Items", "Item", "Item Name", "Product", "Product Name", "Name"]))
        raw_status = clean_text(first_present(row, ["Status", "State"]))
        if not item or not raw_status:
            continue

        quantity = parse_number(first_present(row, ["Reserved", "Qty", "Quantity", "QTY"])) or 0.0
        payout_per_unit = parse_number(first_present(row, ["Payout", "Payout $", "Unit Payout"])) or 0.0
        payout_total = parse_number(first_present(row, ["Subtotal", "Payout Total", "Total Payout", "Subtotal $"]))
        if payout_total is None:
            payout_total = payout_per_unit * quantity
        amount_paid = parse_number(first_present(row, ["Amount Paid", "Paid", "Cash Paid"])) or 0.0
        received = parse_number(first_present(row, ["Received", "Received Qty", "Received Quantity"])) or 0.0
        retail = parse_number(first_present(row, ["Retail Price", "Retail", "Purchase $", "Purchase", "Price"]))
        purchase_total, price_source, price_is_estimate = infer_site_purchase_total(item, retail, quantity, payout_total)

        status = display_status(raw_status)
        profit = calculate_profit(status, payout_total, purchase_total, CASHBACK_RATE)
        order_number = clean_text(first_present(row, ["Order No.", "Order #", "Order", "Order Number"]))
        date_reserved = parse_date(first_present(row, ["Date Reserved", "Reserved At", "Date", "Reserved"]))
        date_processed = parse_date(first_present(row, ["Date Processed", "Processed At", "Processed"]))
        date_paid = parse_date(first_present(row, ["Date Paid", "Paid At"]))
        month_key = date_reserved[:7] if date_reserved else "Unknown"

        records.append(
            {
                "id": len(records) + 1,
                "source_row": index,
                "item_name": item,
                "quantity": quantity,
                "order_number": order_number,
                "tracking": clean_text(first_present(row, ["Tracking", "Tracking #", "Tracking Number"])),
                "insurance": clean_text(first_present(row, ["Insurance"])),
                "payout_per_unit": payout_per_unit,
                "payout_total": round(payout_total, 2),
                "received": received,
                "amount_paid": round(amount_paid, 2),
                "date": date_reserved,
                "date_processed": date_processed,
                "date_paid": date_paid,
                "month_key": month_key,
                "cashback_rate": CASHBACK_RATE,
                "cashback_rate_source": "Default 6% pending Amazon match",
                "purchase_total": round(purchase_total, 2),
                "purchase_is_estimate": price_is_estimate,
                "price_source": price_source,
                "profit": round(profit, 2),
                "account": account_from_order(order_number),
                "account_source": "Waiting for Amazon order history match",
                "amazon_order_matched": False,
                "amazon_profile": "",
                "amazon_payment_method": "",
                "amazon_reward_text": "",
                "status": status,
                "status_raw": raw_status,
            }
        )

    infer_partial_return_details(records)
    infer_split_return_details(records)
    records.sort(key=lambda record: (record["date"] or "", record["source_row"]), reverse=True)
    for index, record in enumerate(records, start=1):
        record["id"] = index

    return {
        "records": records,
        "summary": summarize(records),
        "metadata": {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "tracker_export": "",
            "price_workbook": "",
            "source_url": source_url,
            "source": "bfmr_site",
            "cashback_rate": CASHBACK_RATE,
        },
    }


def normalize_bfmr_export(tracker_export: Path, price_workbook: Path | None = None) -> dict[str, Any]:
    tracker_export = Path(tracker_export)
    price_workbook = Path(price_workbook) if price_workbook else None
    price_lookup = load_price_lookup(price_workbook)

    wb = openpyxl.load_workbook(tracker_export, data_only=True)
    ws = wb.active
    row = find_header_row(ws, {"Status", "Items", "Reserved", "Order No.", "Payout", "Subtotal", "Date Reserved"})
    headers = header_map(ws, row)

    records: list[dict[str, Any]] = []
    for r in range(row + 1, ws.max_row + 1):
        values = {name: cell(ws, headers, r, name) for name in headers}
        item = clean_text(values.get("Items"))
        raw_status = clean_text(values.get("Status"))
        if not item or not raw_status:
            continue

        quantity = parse_number(values.get("Reserved")) or 0.0
        payout_per_unit = parse_number(values.get("Payout")) or 0.0
        payout_total = parse_number(values.get("Subtotal"))
        if payout_total is None:
            payout_total = payout_per_unit * quantity
        amount_paid = parse_number(values.get("Amount Paid")) or 0.0
        received = parse_number(values.get("Received")) or 0.0
        purchase_total, price_source, price_is_estimate = resolve_purchase_total(values, price_lookup)

        status = display_status(raw_status)
        profit = calculate_profit(status, payout_total, purchase_total, CASHBACK_RATE)
        order_number = clean_text(values.get("Order No."))
        date_reserved = parse_date(values.get("Date Reserved"))
        date_processed = parse_date(values.get("Date Processed"))
        date_paid = parse_date(values.get("Date Paid"))
        month_key = date_reserved[:7] if date_reserved else "Unknown"

        records.append(
            {
                "id": len(records) + 1,
                "source_row": r,
                "item_name": item,
                "quantity": quantity,
                "order_number": order_number,
                "tracking": clean_text(values.get("Tracking")),
                "insurance": clean_text(values.get("Insurance")),
                "payout_per_unit": payout_per_unit,
                "payout_total": round(payout_total, 2),
                "received": received,
                "amount_paid": round(amount_paid, 2),
                "date": date_reserved,
                "date_processed": date_processed,
                "date_paid": date_paid,
                "month_key": month_key,
                "cashback_rate": CASHBACK_RATE,
                "cashback_rate_source": "Default 6% pending Amazon match",
                "purchase_total": round(purchase_total, 2),
                "purchase_is_estimate": price_is_estimate,
                "price_source": price_source,
                "profit": round(profit, 2),
                "account": account_from_order(order_number),
                "account_source": "Waiting for Amazon order history match",
                "amazon_order_matched": False,
                "amazon_profile": "",
                "amazon_payment_method": "",
                "amazon_reward_text": "",
                "status": status,
                "status_raw": raw_status,
            }
        )

    infer_partial_return_details(records)
    infer_split_return_details(records)
    records.sort(key=lambda record: (record["date"] or "", record["source_row"]), reverse=True)
    for index, record in enumerate(records, start=1):
        record["id"] = index

    return {
        "records": records,
        "summary": summarize(records),
        "metadata": {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "tracker_export": str(tracker_export),
            "price_workbook": str(price_workbook) if price_workbook else "",
            "cashback_rate": CASHBACK_RATE,
        },
    }


def normalize_gus_tracking_sheet(tracker_sheet: Path) -> dict[str, Any]:
    tracker_sheet = Path(tracker_sheet)
    wb = openpyxl.load_workbook(tracker_sheet, data_only=True)
    ws = wb.active
    row = find_header_row(ws, {"ITEM NAME", "QTY", "ORDER #", "PURCHASE $", "PAYOUT $", "DATE", "STATUS"})
    headers = header_map(ws, row)

    records: list[dict[str, Any]] = []
    for r in range(row + 1, ws.max_row + 1):
        values = {name: cell(ws, headers, r, name) for name in headers}
        item = clean_text(values.get("ITEM NAME"))
        raw_status = clean_text(values.get("STATUS"))
        if not item or not raw_status:
            continue

        quantity = parse_number(values.get("QTY")) or 0.0
        payout_total = parse_number(values.get("PAYOUT $")) or 0.0
        payout_per_unit = payout_total / quantity if quantity else payout_total
        purchase_total = parse_number(values.get("PURCHASE $")) or 0.0
        cashback_rate = parse_number(values.get("CASHBACK %"))
        if cashback_rate is None:
            cashback_rate = CASHBACK_RATE
        if cashback_rate > 1:
            cashback_rate /= 100
        status = display_status(raw_status)
        order_number = clean_text(values.get("ORDER #"))
        date_reserved = parse_date(values.get("DATE"))
        month_key = date_reserved[:7] if date_reserved else "Unknown"
        account = clean_text(values.get("ACCOUNT")) or account_from_order(order_number)
        amount_paid = payout_total if status.lower() == "paid" else 0.0

        records.append(
            {
                "id": len(records) + 1,
                "source_row": r,
                "item_name": item,
                "quantity": quantity,
                "order_number": order_number,
                "tracking": clean_text(values.get("TRACKING #")),
                "insurance": "",
                "payout_per_unit": round(payout_per_unit, 2),
                "payout_total": round(payout_total, 2),
                "received": 0.0,
                "amount_paid": round(amount_paid, 2),
                "date": date_reserved,
                "date_processed": None,
                "date_paid": date_reserved if status.lower() == "paid" else None,
                "month_key": month_key,
                "cashback_rate": round(float(cashback_rate), 4),
                "cashback_rate_source": "Gus tracking sheet",
                "purchase_total": round(purchase_total, 2),
                "purchase_is_estimate": False,
                "price_source": "Gus tracking sheet purchase",
                "profit": round(calculate_profit(status, payout_total, purchase_total, float(cashback_rate)), 2),
                "account": account,
                "account_source": "Gus tracking sheet",
                "amazon_order_matched": False,
                "amazon_profile": "",
                "amazon_payment_method": "",
                "amazon_reward_text": "",
                "amazon_delivery_status": "",
                "amazon_delivery_eta": "",
                "amazon_delivery_eta_date": "",
                "amazon_delivery_scraped_at": "",
                "status": status,
                "status_raw": raw_status,
            }
        )

    infer_partial_return_details(records)
    infer_split_return_details(records)
    records.sort(key=lambda record: (record["date"] or "", record["source_row"]), reverse=True)
    for index, record in enumerate(records, start=1):
        record["id"] = index

    return {
        "records": records,
        "summary": summarize(records),
        "metadata": {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "tracker_export": str(tracker_sheet),
            "price_workbook": "",
            "source": "gus_tracking_sheet",
            "cashback_rate": CASHBACK_RATE,
        },
    }


def summarize(records: list[dict[str, Any]]) -> dict[str, Any]:
    active = [record for record in records if record["status"].lower() != "cancelled"]
    paid = [record for record in active if record["status"].lower() == "paid"]
    missing_tracking = [
        record
        for record in active
        if not record["tracking"] or record["tracking"].lower() == "not submitted"
    ]

    monthly: dict[str, dict[str, Any]] = {}
    for record in records:
        month = record["month_key"] or "Unknown"
        bucket = monthly.setdefault(
            month,
            {
                "month": month,
                "orders": 0,
                "units": 0.0,
                "spend": 0.0,
                "payout": 0.0,
                "profit": 0.0,
                "cash_paid": 0.0,
            },
        )
        bucket["orders"] += 1
        if record["status"].lower() != "cancelled":
            bucket["units"] += record["quantity"]
            bucket["spend"] += record["purchase_total"]
            bucket["payout"] += record["payout_total"]
        bucket["profit"] += record["profit"]
        bucket["cash_paid"] += record["amount_paid"]

    status_counts = Counter(record["status"] for record in records)
    price_counts = Counter(record["price_source"] for record in records)
    account_summary: dict[str, dict[str, Any]] = {}
    item_summary: dict[str, dict[str, Any]] = {}

    for record in records:
        account = account_summary.setdefault(
            record["account"],
            {"account": record["account"], "orders": 0, "spend": 0.0, "payout": 0.0, "profit": 0.0},
        )
        account["orders"] += 1
        if record["status"].lower() != "cancelled":
            account["spend"] += record["purchase_total"]
            account["payout"] += record["payout_total"]
        account["profit"] += record["profit"]

        item = item_summary.setdefault(
            record["item_name"],
            {"item_name": record["item_name"], "orders": 0, "units": 0.0, "spend": 0.0, "profit": 0.0},
        )
        item["orders"] += 1
        item["units"] += record["quantity"]
        if record["status"].lower() != "cancelled":
            item["spend"] += record["purchase_total"]
        item["profit"] += record["profit"]

    def money(value: float) -> float:
        return round(value, 2)

    monthly_rows = [
        {
            **bucket,
            "units": round(bucket["units"], 2),
            "spend": money(bucket["spend"]),
            "payout": money(bucket["payout"]),
            "profit": money(bucket["profit"]),
            "cash_paid": money(bucket["cash_paid"]),
        }
        for bucket in monthly.values()
    ]
    monthly_rows.sort(key=lambda row: row["month"])

    account_rows = []
    for row in account_summary.values():
        account_rows.append(
            {
                **row,
                "spend": money(row["spend"]),
                "payout": money(row["payout"]),
                "profit": money(row["profit"]),
            }
        )
    account_rows.sort(key=lambda row: row["account"])

    top_items = []
    for row in item_summary.values():
        top_items.append(
            {
                **row,
                "spend": money(row["spend"]),
                "profit": money(row["profit"]),
            }
        )
    top_items.sort(key=lambda row: row["profit"], reverse=True)

    return {
        "orders": len(records),
        "active_orders": len(active),
        "paid_orders": len(paid),
        "units": round(sum(record["quantity"] for record in active), 2),
        "spend": money(sum(record["purchase_total"] for record in active)),
        "payout": money(sum(record["payout_total"] for record in active)),
        "profit": money(sum(record["profit"] for record in records)),
        "cash_paid": money(sum(record["amount_paid"] for record in records)),
        "open_payout": money(sum(record["payout_total"] for record in active) - sum(record["amount_paid"] for record in records)),
        "missing_tracking": len(missing_tracking),
        "estimated_purchase_rows": sum(1 for record in records if record["purchase_is_estimate"]),
        "status_counts": dict(status_counts),
        "price_source_counts": dict(price_counts),
        "monthly": monthly_rows,
        "accounts": account_rows,
        "top_items": top_items[:10],
    }


def save_dataset(dataset: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(dataset, indent=2), encoding="utf-8")


def load_dataset(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))
